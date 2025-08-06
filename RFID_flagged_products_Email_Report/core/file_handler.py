import os
from RFID_flagged_products_Email_Report.core.time_utils import get_report_display_date
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from openpyxl.utils.exceptions import IllegalCharacterError
from RFID_flagged_products_Email_Report.config.settings import EXPORT_DIR
from RFID_flagged_products_Email_Report.config.logging_config import configure_logging
from RFID_flagged_products_Email_Report.core.data_processor import remove_timestamp_columns


logger = configure_logging()

def save_to_xlsx(data, query_name):
    """Save the extracted data to an Excel file with proper error handling."""
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)
        logger.info(f"Created directory {EXPORT_DIR}")

    # Create DataFrame
    if not data:
        logger.info(f"No data found for query: {query_name}. Creating empty file.")
        df = pd.DataFrame([{'Message': 'No Data Found'}])
    else:
        df = pd.DataFrame(data)
        df = remove_timestamp_columns(df)

        # Reorder columns if present
        if 'Store' in df.columns and 'Sku' in df.columns:
            df = df[['Store', 'Sku'] + [col for col in df.columns if col not in ['Store', 'Sku']]]

        # Convert numeric columns
        for column in ['Store', 'Sku']:
            if column in df.columns:
                df[column] = pd.to_numeric(df[column], errors='coerce')

    # Prepare Excel file
    filename = f"{query_name}_{get_report_display_date()}.xlsx"
    file_path = os.path.join(EXPORT_DIR, filename)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = query_name[:31]  # Excel sheet name limit

        # Write data
        for row in dataframe_to_rows(df, index=False, header=True):
            try:
                ws.append(row)
            except IllegalCharacterError as e:
                logger.warning(f"Illegal character in row: {e}")
                # Clean the row and try again
                cleaned_row = [str(cell).encode('ascii', 'ignore').decode('ascii') if cell else cell for cell in row]
                ws.append(cleaned_row)

        # Apply number formatting
        num_style = NamedStyle(name="num_style", number_format='0')
        for col_num, column in enumerate(df.columns, 1):
            if pd.api.types.is_numeric_dtype(df[column]):
                for row_num in range(2, len(df) + 2):
                    ws.cell(row=row_num, column=col_num).style = num_style

        # Auto-adjust column widths
        for col_num, column in enumerate(df.columns, 1):
            try:
                max_length = max(df[column].astype(str).map(len).max(), len(column))
                ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = max_length + 2
            except Exception as col_error:
                logger.warning(f"Couldn't adjust width for column {column}: {col_error}")

        # Save workbook
        try:
            wb.save(file_path)
            logger.info(f"Successfully saved Excel file: {file_path}")
            return file_path
        except PermissionError:
            logger.error(f"Permission denied when saving {file_path}")
            return None
        except Exception as save_error:
            logger.error(f"Failed to save Excel file: {save_error}")
            return None

    except Exception as e:
        logger.error(f"Unexpected error creating Excel file: {e}")
        return None