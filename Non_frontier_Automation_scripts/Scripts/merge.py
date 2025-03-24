import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import pytz
from tenacity import retry, stop_after_attempt, wait_fixed
import os
import logging
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
import concurrent.futures
import time

# Timezone for EST
EST = pytz.timezone('US/Eastern')

# Fetch environment variables for secure credentials
API_KEY = os.getenv("NEW_RELIC_API_KEY")
ACCOUNT = os.getenv("NEW_RELIC_ACCOUNT_ID")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
FROM_EMAIL = os.getenv("FROM_EMAIL")
TO_EMAIL = os.getenv("TO_EMAIL")  # Add recipient emails
API_ENDPOINT = 'https://api.newrelic.com/graphql'

# Validate environment variables
if not API_KEY or not ACCOUNT or not EMAIL_PASSWORD or not FROM_EMAIL or not TO_EMAIL:
    logger.error("One or more required environment variables are missing.")
    raise EnvironmentError("Missing required environment variables.")

logger = logging.getLogger('json_logger')
logger.setLevel(logging.INFO)
file_handler = logging.FileHandler('newrelic_data.log')
file_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# Retry decorator
@retry(stop=stop_after_attempt(3), wait=wait_fixed(2))
def fetch_data(nrql_query):
    logger.info(f'Fetching data from New Relic: {nrql_query}')
    headers = {'API-Key': API_KEY, 'Content-Type': 'application/json'}
    query = f"""
    {{
      actor {{
        nrql(
          query:"{nrql_query}"
          accounts: {ACCOUNT}
        ) {{
          results
        }}
      }}
    }}
    """
    payload = {'query': query}
    response = requests.post(API_ENDPOINT, headers=headers, json=payload)
    logger.info(f'Response received. Status Code: {response.status_code}')
    logger.debug(f'Raw Response: {response.text}')
    if 'errors' in response.json():
        logger.error(f"New Relic returned error response: {response.json()['errors'][0]['message']}")
        return []
    response.raise_for_status()
    return response.json()['data']['actor']['nrql']['results']

# Parallel fetching function
def fetch_data_parallel(nrql_query, time_chunks):
    """Fetch data for multiple time chunks in parallel."""
    all_data = []
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = []
        for chunk_start, chunk_end in time_chunks:
            chunk_query = f"{nrql_query} SINCE '{chunk_start.strftime('%Y-%m-%d %H:%M:%S')}' UNTIL '{chunk_end.strftime('%Y-%m-%d %H:%M:%S')}'"
            futures.append(executor.submit(fetch_data, chunk_query))  # Submit the fetch requests in parallel

        for future in concurrent.futures.as_completed(futures):
            chunk_data = future.result()
            all_data.extend(chunk_data)
            logger.info(f"Fetched {len(chunk_data)} events")
    return all_data

def divide_time_range(start_time, end_time, chunk_size_minutes=5):
    """Divide the time range into smaller chunks with adjustable time intervals."""
    time_chunks = []
    current_time = start_time
    while current_time < end_time:
        next_time = current_time + timedelta(minutes=chunk_size_minutes)
        if next_time > end_time:
            next_time = end_time
        time_chunks.append((current_time, next_time))
        current_time = next_time
    return time_chunks

def extract_data(nrql_query, start_time, end_time):
    """Fetch data in chunks for the given time range using parallel fetching."""
    all_data = []
    time_chunks = divide_time_range(start_time, end_time)  # Divide into time chunks
    all_data = fetch_data_parallel(nrql_query, time_chunks)
    return all_data

def remove_timestamp_columns(df):
    """Remove any timestamp-like columns."""
    for column in df.columns:
        if pd.api.types.is_numeric_dtype(df[column]):
            if df[column].max() > 1e12:
                df.drop(columns=[column], inplace=True)
                logger.info(f"Removed column with timestamp data: {column}")
    return df

def save_to_xlsx(data, query_name, local_dir):
    """Save the extracted data to an Excel file."""
    if not os.path.exists(local_dir):
        os.makedirs(local_dir)
        logger.info(f"Created directory {local_dir}")

    if not data:
        logger.info(f"No data found for query: {query_name}. Creating an empty file with a 'No Data Found' message.")
        df = pd.DataFrame([{'Message': 'No Data Found'}])
    else:
        df = pd.DataFrame(data)
        df = remove_timestamp_columns(df)

        # Reorder columns: first 'Store', then 'Sku'
        if 'Store' in df.columns and 'Sku' in df.columns:
            df = df[['Store', 'Sku'] + [col for col in df.columns if col not in ['Store', 'Sku']]]

        for column in ['Store', 'Sku']:
            if column in df.columns:
                df[column] = pd.to_numeric(df[column], errors='coerce')

        for column in df.columns:
            if pd.api.types.is_numeric_dtype(df[column]):
                df[column] = pd.to_numeric(df[column], errors='coerce')

    filename = f"{query_name}_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    file_path = os.path.join(local_dir, filename)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = query_name

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        num_style = NamedStyle(name="num_style", number_format='0')  # No decimal places

        for col_num, column in enumerate(df.columns, 1):
            if column in ['Store', 'Sku'] and pd.api.types.is_numeric_dtype(df[column]):
                for row_num in range(2, len(df) + 2):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.style = num_style

        for col_num, column in enumerate(df.columns, 1):
            max_length = 0
            column = df[column]

            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell)))
                except:
                    pass

            adjusted_width = (max_length + 2)
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = adjusted_width

        wb.save(file_path)
        logger.info(f"Data saved to Excel file: {file_path}")
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        return None

    return file_path

def convert_to_est(dt):
    """Convert a datetime object to Eastern Standard Time (EST)."""
    utc_time = pytz.utc.localize(dt)
    return utc_time.astimezone(EST)

def send_email_with_attachment(subject, body, file_path):
    """Send an email with the Excel file as an attachment."""
    try:
        msg = MIMEMultipart()
        msg['From'] = FROM_EMAIL
        msg['To'] = TO_EMAIL
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        with open(file_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(file_path)}')
            msg.attach(part)

        with smtplib.SMTP('smtp.office365.com', 587) as server:
            server.starttls()
            server.login(FROM_EMAIL, EMAIL_PASSWORD)
            server.sendmail(FROM_EMAIL, TO_EMAIL.split(','), msg.as_string())
            logger.info(f"Email sent to {TO_EMAIL} with attachment {file_path}")
            print(f"Email sent successfully to {TO_EMAIL} with attachment {file_path}")
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        print(f"Failed to send email: {e}")

if __name__ == '__main__':
    # Set the time range for the query
    start_time = datetime(2025, 2, 11, 0, 0, 0)
    end_time = datetime(2025, 2, 12, 23, 59, 59)
    start_time_str = start_time.strftime('%Y-%m-%d %H:%M:%S')
    end_time_str = end_time.strftime('%Y-%m-%d %H:%M:%S')
    logger.info(f"Fetching data from {start_time_str} EST to {end_time_str} EST")

    # List of queries
    queries = [
        {
            "nrql": f"SELECT `Sku`,`Store` FROM Log_RFID WHERE `entity.name` = 'Prod-rfid-cc-results-subscriber' and message LIKE '%[MKS:%] SKU % is written to Local RFID DB%' limit max  ORDER BY timestamp ASC",
            "query_name": "rfid-flagged-skus-mks"
        },
        {
            "nrql": f"SELECT `Sku`,`Store` FROM Log_RFID WHERE `entity.name` = 'Prod-rfid-cc-results-subscriber' AND message LIKE '%[FGL:%] SKU % is written to Local RFID DB%' limit max  ORDER BY timestamp ASC",
            "query_name": "rfid-flagged-skus-fgl"
        }
    ]

    # Loop through each query, extract data, save it to Excel, and send the email

    for query in queries:
        try:
            data = extract_data(query["nrql"], start_time, end_time)
            if not data:
                logger.warning(f"No data found for {query['query_name']}")
                print(f"No data found for {query['query_name']}")
                continue

            file_path = save_to_xlsx(data, query["query_name"], "export")
            if file_path:
                send_email_with_attachment(
                    subject=f"Data for {query['query_name']}",
                    body="Attached is the data you requested.",
                    file_path=file_path
                )
        except Exception as e:
            logger.error(f"Error processing query {query['query_name']}: {e}")
            print(f"Error processing query {query['query_name']}: {e}")

    print("Script execution completed successfully!")
    logger.info("Script execution completed successfully!")