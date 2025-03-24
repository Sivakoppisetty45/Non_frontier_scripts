import pytz
from tenacity import retry, stop_after_attempt, wait_fixed
import os
import logging
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
import concurrent.futures
import time
from dotenv import load_dotenv

# Timezone for EST
EST = pytz.timezone('US/Eastern')

# Load environment variables from .env file
load_dotenv()

API_KEY = os.getenv("NEW_RELIC_API_KEY")
ACCOUNT = os.getenv("NEW_RELIC_ACCOUNT_ID")
API_ENDPOINT = 'https://api.newrelic.com/graphql'

logger = logging.getLogger('json_logger')
logger.setLevel(logging.INFO)
file_handler = logging.FileHandler('newrelic_data.log')
file_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# Retry decorator
@retry(stop=stop_after_attempt(3), wait=wait_fixed(2))
def fetch_data(nrql_query):
    logger.info(f'Fetching data from New Relic: {nrql_query}')
    headers = {'API-Key': API_KEY, 'Content-Type': 'application/json'}
    query = f"""
    {{
      actor {{
        nrql(
          query:"{nrql_query}"
          accounts: {ACCOUNT}
        ) {{
          results
        }}
      }}
    }}
    """
    payload = {'query': query}
    response = requests.post(API_ENDPOINT, headers=headers, json=payload)
    logger.info(f'Response received. Status Code: {response.status_code}')
    logger.debug(f'Raw Response: {response.text}')
    if 'errors' in response.json():
        logger.error(f"New Relic returned error response: {response.json()['errors'][0]['message']}")
        return []
    response.raise_for_status()
    return response.json()['data']['actor']['nrql']['results']

# Parallel fetching function
def fetch_data_parallel(nrql_query, time_chunks):
    """Fetch data for multiple time chunks in parallel."""
    all_data = []
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = []
        for chunk_start, chunk_end in time_chunks:
            chunk_query = f"{nrql_query} SINCE '{chunk_start.strftime('%Y-%m-%d %H:%M:%S')}' UNTIL '{chunk_end.strftime('%Y-%m-%d %H:%M:%S')}'"
            futures.append(executor.submit(fetch_data, chunk_query))  # Submit the fetch requests in parallel

        for future in concurrent.futures.as_completed(futures):
            chunk_data = future.result()
            all_data.extend(chunk_data)
            logger.info(f"Fetched {len(chunk_data)} events")
    return all_data

def divide_time_range(start_time, end_time, chunk_size_minutes=5):
    """Divide the time range into smaller chunks with adjustable time intervals."""
    time_chunks = []
    current_time = start_time
    while current_time < end_time:
        next_time = current_time + timedelta(minutes=chunk_size_minutes)
        if next_time > end_time:
            next_time = end_time
        time_chunks.append((current_time, next_time))
        current_time = next_time
    return time_chunks

def extract_data(nrql_query, start_time, end_time):
    """Fetch data in chunks for the given time range using parallel fetching."""
    all_data = []
    time_chunks = divide_time_range(start_time, end_time)  # Divide into time chunks
    all_data = fetch_data_parallel(nrql_query, time_chunks)
    return all_data

def remove_timestamp_columns(df):
    """Remove any timestamp-like columns."""
    for column in df.columns:
        if pd.api.types.is_numeric_dtype(df[column]):
            if df[column].max() > 1e12:
                df.drop(columns=[column], inplace=True)
                logger.info(f"Removed column with timestamp data: {column}")
    return df

def save_to_xlsx(data, query_name, local_dir):
    """Save the extracted data to an Excel file."""
    if not os.path.exists(local_dir):
        os.makedirs(local_dir)
        logger.info(f"Created directory {local_dir}")

    if not data:
        logger.info(f"No data found for query: {query_name}. Creating an empty file with a 'No Data Found' message.")
        df = pd.DataFrame([{'Message': 'No Data Found'}])
    else:
        df = pd.DataFrame(data)
        df = remove_timestamp_columns(df)

        # Reorder columns: first 'Store', then 'Sku'
        if 'Store' in df.columns and 'Sku' in df.columns:
            df = df[['Store', 'Sku'] + [col for col in df.columns if col not in ['Store', 'Sku']]]

        for column in ['Store', 'Sku']:
            if column in df.columns:
                df[column] = pd.to_numeric(df[column], errors='coerce')

        for column in df.columns:
            if pd.api.types.is_numeric_dtype(df[column]):
                df[column] = pd.to_numeric(df[column], errors='coerce')

    filename = f"{query_name}_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    file_path = os.path.join(local_dir, filename)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = query_name

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        num_style = NamedStyle(name="num_style", number_format='0')  # No decimal places

        for col_num, column in enumerate(df.columns, 1):
            if column in ['Store', 'Sku'] and pd.api.types.is_numeric_dtype(df[column]):
                for row_num in range(2, len(df) + 2):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.style = num_style

        for col_num, column in enumerate(df.columns, 1):
            max_length = 0
            column = df[column]

            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell)))
                except:
                    pass

            adjusted_width = (max_length + 2)
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = adjusted_width

        wb.save(file_path)
        logger.info(f"Data saved to Excel file: {file_path}")
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        return None

    return file_path

def convert_to_est(dt):
    """Convert a datetime object to Eastern Standard Time (EST)."""
    utc_time = pytz.utc.localize(dt)
    return utc_time.astimezone(EST)

def get_yesterday_time_range():
    """Calculate the time range for yesterday (12:00 AM to 11:59 PM)."""
    today = datetime.now(EST).date()
    yesterday = today - timedelta(days=1)

    start_time = datetime.combine(yesterday, datetime.min.time(), tzinfo=EST)  # 12:00 AM
    end_time = datetime.combine(yesterday, datetime.max.time(), tzinfo=EST)  # 11:59 PM

    return start_time, end_time


if __name__ == '__main__':
    # Get time range (from 12:00 AM to 11:59 PM of yesterday)
    start_time, end_time = get_yesterday_time_range()
    start_time_str = start_time.strftime('%Y-%m-%d %H:%M:%S')
    end_time_str = end_time.strftime('%Y-%m-%d %H:%M:%S')

    logger.info(f"Fetching data from {start_time_str} EST to {end_time_str} EST")

    # List of queries
    queries = [
        {
            "nrql": f"SELECT `Sku`,`Store` FROM Log_RFID WHERE `entity.name` = 'Prod-rfid-cc-results-subscriber' and message LIKE '%[MKS:%] SKU % is written to Local RFID DB%' limit max  ORDER BY timestamp ASC",
            "query_name": "rfid-flagged-skus-mks"
        },
        {
            "nrql": f"SELECT `Sku`,`Store` FROM Log_RFID WHERE `entity.name` = 'Prod-rfid-cc-results-subscriber' AND message LIKE '%[FGL:%] SKU % is written to Local RFID DB%' limit max  ORDER BY timestamp ASC",
            "query_name": "rfid-flagged-skus-fgl"
        }
    ]

    # Loop through each query, extract data, and save it to Excel
    for query in queries:
        nrql_query = query["nrql"]
        query_name = query["query_name"]

        # Extract data for the current query
        try:
            data = extract_data(nrql_query, start_time, end_time)
            if not data:
                logger.warning(f"No data found for {query_name}.")
            else:
                print(f"Total events fetched for {query_name}: {len(data)}")

                # Save the data to an Excel file
                file_path = save_to_xlsx(data, query_name, 'export')
                if file_path:
                    logger.info(f"File saved successfully for {query_name} at: {file_path}")
                else:
                    logger.error(f"File saving failed for {query_name}.")
        except Exception as e:
            logger.error(f"Error while fetching data for {query_name}: {str(e)}")

        # Add a small delay between queries to avoid rate limits or timeouts
        time.sleep(2)